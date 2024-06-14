from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from .models import Usuario, Producto, Venta, Categoria, DetalleVenta, Rol, Region, Comuna, Direccion
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# Create your views here.
def inicio(request):
    return render(request, 'core/inicio.html')

def register(request):
    if request.user.is_authenticated:
        return redirect('inicio')
    
    if request.method == "POST":
        nombreUser = request.POST['nombre']
        apellidoUser = request.POST['apellido']
        rutUser = request.POST['rut']
        telefonoUser = request.POST['telefono']
        emailUser = request.POST['email']
        claveUser = request.POST['password']
        confclaveUser = request.POST['confirmPassword']
        
        # Validaciones
        if User.objects.filter(username=emailUser).exists():
            messages.error(request, "Este correo ya está registrado!")
            return redirect('register')
        
        if Usuario.objects.filter(rut=rutUser).exists():
            messages.error(request, "Este RUT ya está registrado!")
            return redirect('register')
        
        if claveUser != confclaveUser:
            messages.error(request, "La contraseña no coincide")
            return redirect('register')
                
        rol = Rol.objects.get(idRol=2)  # Suponiendo que el rol "2" es el rol que deseas asignar
        usuario = Usuario.objects.create(rut=rutUser, nombre=nombreUser, apellido=apellidoUser, telefono=telefonoUser, correo=emailUser, clave=confclaveUser, rol=rol)

        user = User.objects.create_user(username = emailUser, email=emailUser, password= claveUser)
        user.first_name = nombreUser
        user.last_name = apellidoUser
        user.is_staff = True
        user.save()
        messages.success(request, 'Cuenta creada con éxito.')
        
        return redirect('login_user')

    return render(request, 'core/register.html')

def login_user(request):
    if request.user.is_authenticated:
        return redirect('inicio')
    
    if request.method == 'POST':
        email = request.POST['username']
        password = request.POST['pass1']

        user = authenticate(request, username=email, password=password)
        print("Usuario autenticado: ", user)
        if user is not None:
            login(request, user)
            if user.is_superuser:  # Verificar si el usuario es un superusuario
                # Asignar el rol de administrador al superusuario
                rol_administrador = Rol.objects.get(idRol=1)  # Suponiendo que el id del rol de administrador es 1
                user.rol = rol_administrador
            else:
                # Asignar el rol de usuario al usuario normal
                rol_usuario = Rol.objects.get(idRol=2)  # Suponiendo que el id del rol de usuario es 2
                user.rol = rol_usuario
            user.save()
            messages.success(request, 'Inicio de sesión exitoso.')
            return redirect('inicio')
        else:
            messages.error(request, 'Correo o contraseña incorrectos.')

    return render(request, 'core/login_user.html')

def cerrarsesion(request):
    logout(request)
    messages.success(request, "Sesión cerrada correctamente!!")
    return redirect('inicio')


@login_required
def editarperfil(request):
    user = request.user
    usuario = Usuario.objects.get(correo=user.username)
    context = {
        'usuario': usuario
    }
    return render(request, 'core/editarperfil.html', context)

@login_required
def actualizarperfil(request):
    user = request.user
    usuario = Usuario.objects.get(correo=user.username)
    if request.method == "POST":
        nombreUsuario = request.POST['nombre']
        apellidoUsuario = request.POST['apellido']
        telefonoUsuario = request.POST['telefono']
        emailUsuario = request.POST['email']
        
        # Verificar si el correo electrónico ha sido cambiado
        if emailUsuario != user.username:
            # Si el nuevo correo electrónico es diferente al actual del usuario, realizar validación
            if User.objects.filter(username=emailUsuario).exists():
                messages.error(request, "Este correo ya está registrado!")
                return redirect('editarperfil')
        
        usuario.nombre = nombreUsuario
        usuario.apellido = apellidoUsuario
        usuario.telefono = telefonoUsuario
        usuario.correo = emailUsuario
        user.username = emailUsuario
        user.email= emailUsuario
        
        usuario.save()
        user.save()

        messages.success(request, 'Cuenta actualizada con éxito.')
        return redirect('editarperfil')
    
    context = {
        'usuario': usuario
    }
    return render(request, 'core/editarperfil.html', context)


def administrador(request):
    return render(request, 'core/administrador.html')

def agregar(request):
    categoriaProducto = Categoria.objects.all()
    contexto = {
        "categorias" : categoriaProducto
    }

    return render(request, 'core/agregar.html', contexto)

def ingresarproducto(request):

    idProducto = request.POST['id']
    nombreProducto = request.POST['nombre']
    stockProducto = request.POST['stock']
    descripcion = request.POST['descripcion']
    foto = request.FILES['foto']
    precio = request.POST['precio']
    categoria = request.POST['categoria']

    categoriaP= Categoria.objects.get(idCategoria= categoria)

    producto = Producto.objects.create(codProducto= idProducto, nombreP= nombreProducto, stock= stockProducto, descipcion= descripcion, foto= foto, precio= precio, categoria= categoriaP)
    messages.success(request, 'Producto ingresado correctamente.')
    return redirect('agregar')

def listaproducto(request):
    productoListado = Producto.objects.all()
    categorias= Categoria.objects.all()
    contexto = {
        "categorias": categorias,
        "productos": productoListado
    }   
    return render(request, 'core/listaproducto.html', contexto)

def exportar_productos_excel(request):
    productos = Producto.objects.all()
    
    # Crear un libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilo de encabezado
    header_font = Font(bold=True, size=12)
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    # Definir estilo de borde delgado
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Definir colores de fondo para alternar en los títulos
    fill_gray_light = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    fill_gray_dark = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    
    # Desplazar la tabla para que comience en B2
    start_col = 2  # Comenzar en la columna B
    start_row = 2  # Comenzar en la fila 2
    
    # Definir encabezados
    headers = ["ID Producto", "Nombre", "Stock", "Precio", "Categoría"]
    for col_num, header in enumerate(headers, start_col):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border  # Aplicar borde al encabezado
        # Alternar colores
        cell.fill = fill_gray_light if col_num % 2 == 0 else fill_gray_dark
    
    # Añadir datos de productos
    for row_num, producto in enumerate(productos, start_row + 1):
        cells = [
            ws.cell(row=row_num, column=start_col, value=producto.codProducto),
            ws.cell(row=row_num, column=start_col + 1, value=producto.nombreP),
            ws.cell(row=row_num, column=start_col + 2, value=producto.stock),
            ws.cell(row=row_num, column=start_col + 3, value=producto.precio),
            ws.cell(row=row_num, column=start_col + 4, value=producto.categoria.nombreCa if hasattr(producto.categoria, 'nombreCa') else '')
        ]
        for cell in cells:
            cell.alignment = alignment_center
            cell.border = thin_border  # Aplicar borde a las celdas de datos
            cell.font = Font(size=12)  # Establecer tamaño de fuente en 12
    
    # Ajustar el ancho de las columnas basado en el contenido
    for col in ws.iter_cols(min_col=start_col, max_col=start_col + len(headers) - 1):
        max_length = 0
        column = col[0].column_letter  # Obtener el nombre de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajuste adicional del ancho
        ws.column_dimensions[column].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

def buscar_productos(request):
    query = request.GET.get('q')
    if query:
        productos = Producto.objects.filter(nombreP__icontains=query)
    else:
        productos = Producto.objects.all()
    
    context = {
        'productos': productos
    }
    return render(request, 'core/listaproducto.html', context)

def editarproducto(request, id_producto):
    producto = Producto.objects.get(codProducto = id_producto)
    listaCategoria = Categoria.objects.all()
    contexto = {
        "producto" : producto,
        "listacategoria" : listaCategoria
    }
    return render(request, 'core/editarproducto.html', contexto)

def actualizaproducto(request):
    if request.method == "POST":
        idProducto = request.POST['id']
        nombreProducto = request.POST['nombre']
        stockProducto = request.POST['stock']
        descripcion = request.POST['descripcion']
        precio = request.POST['precio']
        categoria = request.POST['categoria']

        producto = Producto.objects.get(codProducto=idProducto)
        categoriaP = Categoria.objects.get(idCategoria=categoria)

        producto.nombreP = nombreProducto
        producto.stock = stockProducto
        producto.descipcion = descripcion
        producto.precio = precio
        producto.categoria = categoriaP

        if 'imagen' in request.FILES:
            producto.foto = request.FILES['imagen']

        producto.save()
        return redirect('listaproducto')


def borrarproducto(request, id_producto):
    productoborrar = Producto.objects.get(codProducto = id_producto)
    productoborrar.delete()
    messages.success(request, 'Producto eliminado correctamente.')
    return redirect('listaproducto')


def listausuarios(request):
    usuariosListado = Usuario.objects.exclude(rut=None).exclude(rut='').all()
    contexto = {
        "usuarios": usuariosListado
    }
    return render(request, 'core/listausuarios.html', contexto)

def borrarperfil(request, rut):
    perfilborrar = Usuario.objects.get(rut = rut)
    perfilborrar.delete()
    messages.success(request, 'Perfil eliminado correctamente.')
    return redirect('listausuarios')

@login_required 
def productos(request):
    productoListado = Producto.objects.all()
    contexto = {
        "productos": productoListado
    }   
    return render(request, 'core/productos.html', contexto)

def detalleproducto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render (request, 'core/detalleproducto.html', {'producto': producto})

def quienessomos(request):
    return render(request, 'core/quienessomos.html')

def galeria(request):
    return render(request, 'core/galeria.html')











